import json
from unicodedata import decimal
from django.http import HttpResponse, JsonResponse
from calendar import day_name
from pyexpat.errors import messages
from django.db import transaction

from webbrowser import get
from django.shortcuts import render, get_object_or_404, redirect
from .forms import ReservationForm, KhoanThuForm, DotThuPhiForm
from .models import HoKhau, NhanKhau, TaiKhoan, VaiTro, KhoanThu, DotThuPhi, HoaDon, BienDongNhanKhau
from datetime import datetime, timezone
from django.contrib import messages
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.db.models import Q,Sum, Count
from django.contrib.auth.hashers import make_password
from django.db.models.functions import ExtractMonth
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.utils import timezone 
# views.py
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
def home(request):
    HoKhaus = HoKhau.objects.all()
    return render(request, 'core/mainpage.html')

def user_logout(request):
    logout(request)
#    messages.success(request, "B·∫°n ƒë√£ ƒëƒÉng xu·∫•t.")
    return redirect("home")

@login_required(login_url="login")

def profile(request):
    print("adsgdgd")
    user= request.user
    print(user.is_authenticated)
    if user.is_authenticated:
        id_vaitro= user.vaitro.id_vaitro
        if id_vaitro is not None:
            return render(request, "core/profile.html", {
                "user": user
                })
    return render(request, "core/message.html", {
        "error": "B·∫°n ch∆∞a ƒëƒÉng nh·∫≠p"
    })

# def register(request):
#     if request.method == "POST":
#         username = request.POST.get("username")
#         password1 = request.POST.get("password1")
#         password2 = request.POST.get("password2")
#         vaitro_string = request.POST.get("vaitro")
#         match vaitro_string: 
#             case "admin":
#                 vaitro=get_object_or_404(VaiTro, id_vaitro=1)
#         match vaitro_string: 
#             case "user":
#                 vaitro=get_object_or_404(VaiTro, id_vaitro=2)
#         match vaitro_string: 
#             case "ketoan":
#                 vaitro=get_object_or_404(VaiTro, id_vaitro=3)
        
#         # 1. Ki·ªÉm tra d·ªØ li·ªáu
#         if not username or not password1 or not password2:
#             return render(request, "core/register.html", {
#                 "error": "Vui l√≤ng nh·∫≠p ƒë·∫ßy ƒë·ªß th√¥ng tin."
#             })

#         if password1 != password2:
#             return render(request, "core/register.html", {
#                 "error": "M·∫≠t kh·∫©u kh√¥ng kh·ªõp"
#             })

#         if TaiKhoan.objects.filter(username=username).exists():
#             return render(request, "core/register.html", {
#                 "error": "T√™n ƒëƒÉng nh·∫≠p ƒë√£ t·ªìn t·∫°i "
#             })

#         # 2. T·∫°o t√†i kho·∫£n
#         taikhoan = TaiKhoan.objects.create(
#             username=username,
#             password=make_password(password1),  # üîê m√£ h√≥a m·∫≠t kh·∫©u
#             vaitro=vaitro,
#             is_active=True,
#             is_staff=False
#         )

#         # 3. T·ª± ƒë·ªông ƒëƒÉng nh·∫≠p sau khi ƒëƒÉng k√Ω
#         login(request, taikhoan)

#         messages.success(request, "ƒêƒÉng k√Ω th√†nh c√¥ng!")
#         return redirect("home")



def login_view(request):
    print("abdcsfds")
    if request.method == "POST":
        username_request = request.POST.get("username")
        password_request = request.POST.get("password")

        user = authenticate(request, username=username_request, password=password_request)
        print(username_request+'\n'+password_request)
        if user is not None:
            login(request, user)
            id_vaitro= request.user.vaitro.id_vaitro
            if id_vaitro is not None:
                if id_vaitro == 1 or id_vaitro==2:
                    return redirect("admin_home")
                if id_vaitro == 3:
                    return redirect("accountant_home")
        else:
            return render(request, "core/login.html", {
                "error": "Sai t√™n ƒëƒÉng nh·∫≠p ho·∫∑c m·∫≠t kh·∫©u"
            })

    return render(request, "core/login.html")


# ========================================================
# 1. H√ÄM LOGIN & HOME (PHI√äN B·∫¢N CH√çNH TH·ª®C - ƒê√É FIX L·ªñI)
# ========================================================

# def login(request):
#     # N·∫øu ƒë√£ c√≥ session th√¨ v√†o th·∫≥ng trang ch·ªß, kh√¥ng c·∫ßn ƒëƒÉng nh·∫≠p l·∫°i
#     if request.session.get('id_taikhoan'):
#         return redirect('home')

#     if request.method == 'POST':
#         # L·∫•y d·ªØ li·ªáu t·ª´ form (t√™n input ƒë√£ kh·ªõp v·ªõi file HTML m·ªõi)
#         user_nhap = None
#         if 'username_admin' in request.POST:
#             user_nhap = request.POST.get('username_admin')
#             pass_nhap = request.POST.get('password_admin')
#             vaitro_mong_muon = 1
#         elif 'username_user' in request.POST:
#             user_nhap = request.POST.get('username_user')
#             pass_nhap = request.POST.get('password_user')
#             vaitro_mong_muon = 3
        
#         if user_nhap:
#             try:
#                 tai_khoan = TaiKhoan.objects.get(username=user_nhap)
                
#                 # So s√°nh m·∫≠t kh·∫©u (Kh√¥ng m√£ h√≥a: 2005 == 2005)
#                 if pass_nhap == tai_khoan.password:
                    
#                     # Ki·ªÉm tra vai tr√≤
#                     real_role = getattr(tai_khoan, 'id_vaitro_id', None)
#                     if real_role is None: 
#                         real_role = getattr(tai_khoan, 'vaitro_id', None)
                    
#                     if real_role == vaitro_mong_muon:
#                         # === L∆ØU SESSION ===
#                         request.session['id_taikhoan'] = tai_khoan.id_taikhoan
#                         # C√°c d√≤ng n√†y ƒë·∫£m b·∫£o Session ƒë∆∞·ª£c l∆∞u ch·∫∑t ch·∫Ω
#                         request.session.modified = True 
#                         request.session.save()
                        
#                         # === CHUY·ªÇN H∆Ø·ªöNG V·ªÄ TRANG CH·ª¶ ===
#                         if real_role == 3:
#                             return redirect('accountant_home')
#                         else:
#                             return redirect('home')
#                     else:
#                         messages.error(request, "B·∫°n ƒëang ƒëƒÉng nh·∫≠p sai vai tr√≤!")
#                 else:
#                     messages.error(request, "M·∫≠t kh·∫©u kh√¥ng ƒë√∫ng!")

#             except TaiKhoan.DoesNotExist:
#                 messages.error(request, "T√™n ƒëƒÉng nh·∫≠p kh√¥ng t·ªìn t·∫°i!")
        
#     return render(request, 'core/login.html')

def admin_home(request):
    # # Ki·ªÉm tra b·∫£o m·∫≠t: Ch∆∞a ƒëƒÉng nh·∫≠p th√¨ ƒëu·ªïi v·ªÅ Login
    # id_tk = request.session.get('id_taikhoan')
    # if not id_tk:
    #     return redirect('login')
        
    # ƒê√£ ƒëƒÉng nh·∫≠p -> Hi·ªÉn th·ªã trang ch·ªß
    HoKhaus = HoKhau.objects.all()
    return render(request, 'core/admin_home.html', {'admin_home': HoKhaus})
@login_required(login_url="login")

def accountant_home(request):
    # Ki·ªÉm tra b·∫£o m·∫≠t

        
    return render(request, 'core/Accountant.html')

# ========================================================
# 2. C√ÅC H√ÄM CH·ª®C NƒÇNG KH√ÅC (GI·ªÆ NGUY√äN)
# ========================================================

def test(request):
    ds_ho_khau = HoKhau.objects.all()
    return render(request, 'core/test.html', {'ho_khau_list': ds_ho_khau})
@login_required(login_url="login")

def edit_nhan_khau(request, id_nhankhau):
    nk = get_object_or_404(NhanKhau, id_nhankhau=id_nhankhau)
    if request.method == "POST":
        nk.ho_ten = request.POST.get('ho_ten')
        nk.ngay_sinh = request.POST.get('ngay_sinh') or None
        nk.cccd = request.POST.get('cccd') or None
        nk.quan_he_chu_ho = request.POST.get('quan_he_chu_ho') or None
        ho_khau_id = request.POST.get('ho_khau_id')

        if ho_khau_id:
            nk.id_hokhau_id = ho_khau_id

        nk.save()
        return redirect('nhan_khau_profile', id_nhankhau=nk.id_nhankhau)

    return render(request, 'core/demomanage_edit.html', {'nhan_khau': nk})
@login_required(login_url="login")
def nhan_khau_profile(request, id_nhankhau):
    nhan_khau = get_object_or_404(NhanKhau, id_nhankhau=id_nhankhau)
    return render(request, 'core/nhan_khau_profile.html', {'nhan_khau': nhan_khau})
@login_required(login_url="login")

def nhan_khau_delete(request, id_nhankhau):
    exists = NhanKhau.objects.filter(id_nhankhau=id_nhankhau).exists()
    if(exists):
        nhan_khau = get_object_or_404(NhanKhau, id_nhankhau=id_nhankhau)
        nhan_khau.is_deleted=True
    return render(request, 'core/demomanage_delete.html')
@login_required(login_url="login")

def add_demo(request):
    if request.method == "POST":
        ho_ten = request.POST.get('ho_ten')
        ngay_sinh = request.POST.get('ngay_sinh')
        cccd = request.POST.get('cccd')
        quan_he_chu_ho = request.POST.get('quan_he_chu_ho')
        ho_khau_id = request.POST.get('ho_khau_id')
        loai_bien_dong_query = request.POST.get('loai_dang_ky_cu_tru')
        try:
            if HoKhau.objects.filter(id_hokhau=ho_khau_id).exists():
                nhan_khau = NhanKhau.objects.create(
                    ho_ten=ho_ten,
                    ngay_sinh=ngay_sinh or None,
                    cccd=cccd or None,
                    quan_he_chu_ho=quan_he_chu_ho or None,
                    id_hokhau_id=int(ho_khau_id)
                )
                BienDongNhanKhau.objects.create(
                    loai_biendong="thuong tru",   
                    ngay_batdau=timezone.now().date(),
                    id_nhankhau=nhan_khau,
                    ly_do="Dang ky nhan khau moi"
                )
        except HoKhau.DoesNotExist:
            pass 

        return redirect('demomanage/adddemo')   

    return render(request, 'core/demomanage_add.html')

# def login(request):
#     # X·ª≠ l√Ω khi ng∆∞·ªùi d√πng NH·∫§N N√öT (g·ª≠i form)
#     if request.method == 'POST':
        
#         # X√°c ƒë·ªãnh xem form Admin hay User ƒë∆∞·ª£c g·ª≠i
#         if 'username_admin' in request.POST:
#             user_nhap = request.POST.get('username_admin')
#             pass_nhap = request.POST.get('password_admin')
#             # ƒê·∫£m b·∫£o t√™n vai tr√≤ n√†y KH·ªöP CH√çNH X√ÅC v·ªõi CSDL c·ªßa b·∫°n
#             vaitro_mong_muon = 1 # (1 = admin)
        
#         elif 'username_user' in request.POST:
#             user_nhap = request.POST.get('username_user')
#             pass_nhap = request.POST.get('password_user')
#             # ƒê·∫£m b·∫£o t√™n vai tr√≤ n√†y KH·ªöP CH√çNH X√ÅC v·ªõi CSDL c·ªßa b·∫°n
#             vaitro_mong_muon = 3 # (3 = k·∫ø to√°n, ho·∫∑c vai tr√≤ ng∆∞·ªùi d√πng)
        
#         else:
#             user_nhap = None

#         if user_nhap:
#             try:
#                 # B∆∞·ªõc 2: T√¨m t√†i kho·∫£n trong CSDL
#                 tai_khoan = TaiKhoan.objects.get(username=user_nhap) 
                
#                 # B∆∞·ªõc 3: Ki·ªÉm tra m·∫≠t kh·∫©u
#                 if pass_nhap == tai_khoan.password:
                    
#                     # === S·ª¨A L·∫†I D√íNG N√ÄY ===
#                     # B∆∞·ªõc 4: Ki·ªÉm tra vai tr√≤ (So s√°nh S·ªê v·ªõi S·ªê)
#                     if tai_khoan.vaitro_id == vaitro_mong_muon:
#                     # === K·∫æT TH√öC S·ª¨A ===
                        
#                         # B∆Ø·ªöC 5: ƒêƒÇNG NH·∫¨P
#                         # D√πng h√†m 'auth_login' ch√∫ng ta ƒë√£ import
#                         request.session['id_taikhoan'] = tai_khoan.id_taikhoan
                        
#                         # D·ª±a theo th√¥ng tin tr∆∞·ªõc ƒë√≥, 1=Admin, 3=K·∫ø to√°n
                        
#                         if vaitro_mong_muon == 3: # N·∫æU L√Ä K·∫æ TO√ÅN
#                             return redirect('accountant_home') # <-- ƒêi ƒë·∫øn trang K·∫ø to√°n
#                         else: # N·∫æU L√Ä ADMIN HO·∫∂C VAI TR√í KH√ÅC
#                             return redirect('home') # <-- ƒêi ƒë·∫øn trang ch·ªß chung
#                     else:
#                         # Vai tr√≤ sai
#                         messages.error(request, "B·∫°n ƒëang ƒëƒÉng nh·∫≠p ·ªü form kh√¥ng ƒë√∫ng vai tr√≤!")
#                 else:
#                     # M·∫≠t kh·∫©u sai
#                     messages.error(request, "M·∫≠t kh·∫©u kh√¥ng ƒë√∫ng!")

#             except TaiKhoan.DoesNotExist:
#                 # Kh√¥ng t√¨m th·∫•y username
#                 messages.error(request, "T√™n ƒëƒÉng nh·∫≠p kh√¥ng t·ªìn t·∫°i!")
        
#         # N·∫øu c√≥ b·∫•t k·ª≥ l·ªói n√†o, render l·∫°i trang login
#         # (Template s·∫Ω t·ª± ƒë·ªông hi·ªÉn th·ªã c√°c 'messages' l·ªói)
#         return render(request, 'core/login.html')

#     else:
#         # X·ª≠ l√Ω khi ng∆∞·ªùi d√πng M·ªû TRANG (y√™u c·∫ßu GET)
#         # Ch·ªâ c·∫ßn hi·ªÉn th·ªã trang login
#         return render(request, 'core/login.html')
@login_required(login_url="login")

def demomanage(request):
    nhan_khau_list = NhanKhau.objects.all()
    query = request.GET.get('search_id', '')
    if query:
        try:
            for a in nhan_khau_list:
                if(a.id_nhankhau==int(query)):
                    nhan_khau_list= [a]
        except ValueError:
            nhan_khau_list = NhanKhau.objects.all()

    data = []
    for nk in nhan_khau_list:
        bien_dong = (
            BienDongNhanKhau.objects
            .filter(id_nhankhau=nk)
            .order_by('-ngay_batdau')
            .first()
        )

        trang_thai = bien_dong.loai_biendong if bien_dong else "Ch∆∞a x√°c ƒë·ªãnh"

        data.append({
            'nhan_khau': nk,
            'trang_thai': trang_thai
        })
    context = {
        'data': data,
        'query': query,
    }
    return render(request, 'core/demomanage.html', context)

@login_required(login_url="login")
def biendong_list(request):
    biendongs = (
        BienDongNhanKhau.objects
        .select_related('id_nhankhau')
        .order_by('-ngay_batdau')
    )

    return render(
        request,
        'core/biendong_list.html',
        {'biendongs': biendongs}
    )
@login_required(login_url="login")
def add_tam_vang(request, id_nhankhau):
    nhan_khau = get_object_or_404(NhanKhau, id_nhankhau=id_nhankhau)

    if request.method == "POST":
        ngay_batdau = request.POST.get('ngay_batdau')
        ngay_ketthuc = request.POST.get('ngay_ketthuc')
        ly_do = request.POST.get('ly_do')

        with transaction.atomic():
            BienDongNhanKhau.objects.create(
                loai_biendong="tam vang",      # üîí c·ªë ƒë·ªãnh
                ngay_batdau=ngay_batdau or timezone.now().date(),
                ngay_ketthuc=ngay_ketthuc or None,
                ly_do=ly_do or None,
                id_nhankhau=nhan_khau
            )

        return redirect('demomanage/adddemo')  # ho·∫∑c trang chi ti·∫øt nh√¢n kh·∫©u

    return render(
        request,
        'core/add_tam_vang.html',
        {'nhan_khau': nhan_khau}
    )
@login_required(login_url="login")

def demomanage_delete(request, id_hokhau):
    exists = HoKhau.objects.filter(id_hokhau=id_hokhau).exists()
    if(exists):
        hr = get_object_or_404(HoKhau, id_hokhau=id_hokhau)
        hr.is_deleted=True
    return render(request, 'core/hrmanage_delete.html')
@login_required(login_url="login")

def accountmanage(request):
    tai_khoan_list = TaiKhoan.objects.all()
    query = request.GET.get('search_id', '')
    user = request.user
    print(user.is_authenticated)
    if user.is_authenticated:
        id_vaitro= request.user.vaitro.id_vaitro
        if id_vaitro is not None:
            if id_vaitro == 1:
                if query:
                    try:
                        for a in tai_khoan_list:
                            if(a.id_taikhoan==int(query)):
                                tai_khoan_list= [a]
                    except ValueError:
                        tai_khoan_list= TaiKhoan.objects.all()
                
                context = {
                    'tai_khoan_list': tai_khoan_list,
                    'query': query,
                }

                return render(request, 'core/accountmanage.html', context)
    return render(request, "core/message.html", {
        "error": "B·∫°n kh√¥ng c√≥ quy·ªÅn truy c·∫≠p trang n√†y"
    })
@login_required(login_url="login")

def accountmanage_delete(request, id_taikhoan):
    exists = TaiKhoan.objects.filter(id_taikhoan=id_taikhoan).exists()
    if(exists):
        account = get_object_or_404(TaiKhoan, id_taikhoan=id_taikhoan)
        account.is_deleted=True
    return render(request, 'core/accountmanage_delete.html')
@login_required(login_url="login")

def accountmanage_addaccount(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password1 = request.POST.get("password1")
        password2 = request.POST.get("password2")
        vaitro_string = request.POST.get("vaitro")
        match vaitro_string: 
            case "admin":
                vaitro=get_object_or_404(VaiTro, id_vaitro=1)
        match vaitro_string: 
            case "user":
                vaitro=get_object_or_404(VaiTro, id_vaitro=2)
        match vaitro_string: 
            case "ketoan":
                vaitro=get_object_or_404(VaiTro, id_vaitro=3)
        
        # 1. Ki·ªÉm tra d·ªØ li·ªáu
        if not username or not password1 or not password2:
            return render(request, "core/register.html", {
                "error": "Vui l√≤ng nh·∫≠p ƒë·∫ßy ƒë·ªß th√¥ng tin."
            })

        if password1 != password2:
            return render(request, "core/register.html", {
                "error": "M·∫≠t kh·∫©u kh√¥ng kh·ªõp"
            })

        if TaiKhoan.objects.filter(username=username).exists():
            return render(request, "core/register.html", {
                "error": "T√™n ƒëƒÉng nh·∫≠p ƒë√£ t·ªìn t·∫°i "
            })

        # 2. T·∫°o t√†i kho·∫£n
        taikhoan = TaiKhoan.objects.create(
            username=username,
            password=make_password(password1),  # üîê m√£ h√≥a m·∫≠t kh·∫©u
            vaitro=vaitro,
            is_active=True,
            is_staff=False
        )

    return render(request, 'core/accountmanage_addaccount.html', {
                "error": "T·∫°o t√†i kho·∫£n th√†nh c√¥ng "
            })
@login_required(login_url="login")

def view_taikhoan(request, id_taikhoan):

    user = authenticate(request, username="admin", password="2005")
    login(request, user)
    
    taikhoan = get_object_or_404(TaiKhoan, id_taikhoan=id_taikhoan)
    return render(request, 'core/accountmanage_view.html', {'taikhoan': taikhoan})
@login_required(login_url="login")

def edit_taikhoan(request, id_taikhoan):
    taikhoan = get_object_or_404(TaiKhoan, id_taikhoan=id_taikhoan)

    if request.method == "POST":
        username = request.POST.get('username')
        password_raw = request.POST.get('password')
        vaitro_id = request.POST.get('vaitro_id')
        try:
            vaitro = VaiTro.objects.get(id_vaitro=vaitro_id)
            taikhoan.vaitro = vaitro
        except VaiTro.DoesNotExist:
            return render(request, "core/message.html", {
                "error": "ID vai tr√≤ kh√¥ng t·ªìn t·∫°i"
            })
        if not taikhoan.username==username:
            if not TaiKhoan.objects.filter(username=username).exists():
                taikhoan.username=username
            else:
                return render(request, "core/message.html", {
                    "error": "username n√†y ƒë√£ t·ªìn t·∫°i"
                })

        taikhoan.set_password(password_raw)

        taikhoan.save()
        return render(request, "core/message.html", {
            "error": "Thay ƒë·ªïi th√¥ng tin th√†nh c√¥ng"
        })
    return render(request, 'core/accountmanage_change.html', {'taikhoan': taikhoan})

@login_required(login_url="login")

def hrmanage(request):
    print(2)
    ds_ho_khau = HoKhau.objects.all()
    query = request.GET.get('search_id', '')
    if query:
        try:
            for a in ds_ho_khau:
                if(a.id_hokhau==int(query)):
                    ds_ho_khau= [a]
        except ValueError:
            ds_ho_khau = HoKhau.objects.all()
    
    context = {
        'ho_khau_list': ds_ho_khau,
        'query': query,
    }

    return render(request, 'core/hrmanage.html', context)
@login_required(login_url="login")

def hrmanage_delete(request, id_hokhau):
    exists = HoKhau.objects.filter(id_hokhau=id_hokhau).exists()
    if(exists):
        hr = get_object_or_404(HoKhau, id_hokhau=id_hokhau)
        hr.is_deleted=True
    return render(request, 'core/hrmanage_delete.html')
@login_required(login_url="login")

def add_hokhau(request):
    if request.method == "POST":
        so_can_ho = request.POST.get('so_can_ho')
        dien_tich = request.POST.get('dien_tich')

        if so_can_ho:
            try:
                dien_tich_value = float(dien_tich) if dien_tich else None
                HoKhau.objects.create(so_can_ho=so_can_ho, dien_tich=dien_tich_value)
                messages.success(request, f"H·ªô kh·∫©u cƒÉn {so_can_ho} ƒë√£ ƒë∆∞·ª£c th√™m th√†nh c√¥ng!")
            except ValueError:
                messages.error(request, "Gi√° tr·ªã di·ªán t√≠ch kh√¥ng h·ª£p l·ªá!")
        else:
            messages.error(request, "Vui l√≤ng nh·∫≠p s·ªë cƒÉn h·ªô!")

        return redirect('add_hokhau')

    return render(request, 'core/add_hokhau.html')
@login_required(login_url="login")

def hokhau_detail(request, id_hokhau):
    print(3)
    hokhau = HoKhau.objects.get( id_hokhau=id_hokhau)
    print(hokhau.id_hokhau )
    thanh_vien = NhanKhau.objects.filter(id_hokhau_id=id_hokhau)

    return render(request, 'core/hokhau_detail.html', {
        'hokhau': hokhau,
        'thanh_vien': thanh_vien
    })

@login_required(login_url="login")

def edit_hokhau(request, id_hokhau):
    hokhau = get_object_or_404(HoKhau, id_hokhau=id_hokhau)

    if request.method == "POST":
        so_can_ho = request.POST.get("so_can_ho")
        dien_tich = request.POST.get("dien_tich")

        if not so_can_ho:
            messages.error(request, "S·ªë cƒÉn h·ªô kh√¥ng ƒë∆∞·ª£c ƒë·ªÉ tr·ªëng!")
        else:
            try:
                hokhau.so_can_ho = so_can_ho
                hokhau.dien_tich = float(dien_tich) if dien_tich else None
                hokhau.save()
                messages.success(request, "C·∫≠p nh·∫≠t th√¥ng tin h·ªô kh·∫©u th√†nh c√¥ng!")
                return redirect('hrmanage')
            except ValueError:
                messages.error(request, "Di·ªán t√≠ch ph·∫£i l√† s·ªë h·ª£p l·ªá!")

    return render(request, 'core/hokhau_edit.html', {'hokhau': hokhau})
@login_required(login_url="login")

def HoKhaus_list(request):
    HoKhaus = HoKhau.objects.all()
    return render(request, 'core/HoKhaus_list.html', {'HoKhaus': HoKhaus})


@login_required(login_url="login")

#============= K·∫ø to√°n Views =================
def accountant_home(request):
    return render(request, 'core/Accountant.html')
#Quan l√Ω kho·∫£n thu
def fee_management(request):
    query = request.GET.get('search_khoanthu') #L·∫•y tham s·ªë t√¨m ki·∫øm t·ª´ URL
    khoan_thu_list = KhoanThu.objects.all() # L·∫•y d·ªØ li·ªáu t·ª´ database
    
    if query:
        khoan_thu_list = khoan_thu_list.filter(
            Q(ten_khoanthu__icontains=query) | Q(id_khoanthu__icontains=query)
        )
    total_count = khoan_thu_list.count()
    form = KhoanThuForm()
    context = {
        'khoan_thu_list': khoan_thu_list, 
        'total_count': total_count,      
        'form': form, 
        'query': query
    }
    return render(request, 'core/FeeManagement.html', context)
@login_required(login_url="login")

def view_khoanthu_detail_modal(request, pk):
    khoan_thu = get_object_or_404(KhoanThu, id_khoanthu=pk) 
    if not request.headers.get('x-requested-with') == 'XMLHttpRequest':
        pass 
    return render(request, 'core/ViewFeeDetailModal.html', {'khoan_thu': khoan_thu})
@login_required(login_url="login")

def add_khoanthu(request):
    if request.method == 'POST':
        form = KhoanThuForm(request.POST)
        if form.is_valid():
            new_ID = form.cleaned_data.get('id_khoanthu')
            # Ki·ªÉm tra tr√πng m√£ kho·∫£n thu
            if KhoanThu.objects.filter(id_khoanthu=new_ID).exists():
                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({'error': 'M√£ kho·∫£n thu n√†y ƒë√£ t·ªìn t·∫°i!'}, status=400)
                messages.error(request, 'M√£ kho·∫£n thu n√†y ƒë√£ t·ªìn t·∫°i!')
                return render(request, 'core/AddFeeModal.html', {'form': form})
            
            form.save()

            # N·∫øu l√† y√™u c·∫ßu t·ª´ Modal (AJAX)
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'success'})
            
            # N·∫øu l√† y√™u c·∫ßu th√¥ng th∆∞·ªùng, quay l·∫°i trang qu·∫£n l√Ω kho·∫£n thu
            messages.success(request, 'Th√™m kho·∫£n thu th√†nh c√¥ng!')
            return redirect('fee_management')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'error': 'D·ªØ li·ªáu kh√¥ng h·ª£p l·ªá!'}, status=400)
    else:
        form = KhoanThuForm()
    
    return render(request, 'core/AddFeeModal.html', {'form': form})
@login_required(login_url="login")
def edit_khoanthu(request, pk):
    khoan_thu = get_object_or_404(KhoanThu, id_khoanthu=pk) 

    if request.method == 'POST':
        form = KhoanThuForm(request.POST, instance=khoan_thu)
        
        if form.is_valid():
            form.save() 
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'success'})
            messages.success(request, f"Kho·∫£n thu '{khoan_thu.ten_khoanthu}' ƒë√£ ƒë∆∞·ª£c c·∫≠p nh·∫≠t th√†nh c√¥ng!")
            return redirect('fee_management')
        else:
            context = {'form': form, 'khoan_thu': khoan_thu}
            return render(request, 'core/EditFeeModal.html', context, status=400)
            
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        form = KhoanThuForm(instance=khoan_thu)
        context = {'form': form, 'khoan_thu': khoan_thu}
        return render(request, 'core/EditFeeModal.html', context) 
    else:
        form = KhoanThuForm(instance=khoan_thu)
        context = {
            'form': form, 
            'khoan_thu': khoan_thu,
            'page_title': f"CH·ªàNH S·ª¨A KHO·∫¢N THU - {khoan_thu.ten_khoanthu}",
        }
        return render(request, 'core/EditFee.html', context)
@login_required(login_url="login")

def delete_khoanthu(request, pk):
    khoan_thu = get_object_or_404(KhoanThu, id_khoanthu=pk)

    if request.method == 'POST':
        khoan_thu.delete()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success', 'message': 'Kho·∫£n thu ƒë√£ ƒë∆∞·ª£c x√≥a th√†nh c√¥ng!'})
        return redirect('fee_management')
    return render(request, 'core/DeleteFeeModal.html', {'khoan_thu': khoan_thu})
@login_required(login_url="login")
    
#Qu·∫£n l√Ω ƒë·ª£t thu ph√≠
def fee_collection_period(request):
    query = request.GET.get('search_dotthu')
    dot_thu_phi_list = DotThuPhi.objects.select_related('id_khoanthu').all()
    active_count = dot_thu_phi_list.filter(trang_thai='open').count()
    if query:
        dot_thu_phi_list = dot_thu_phi_list.filter(
            Q(ten_dotthu__icontains=query) | Q(id_dotthu__icontains=query)
        )
    context = {
        'dot_thu_phi_list': dot_thu_phi_list,
        'active_count': active_count,
        'query': query
    }
    return render(request, 'core/FeeCollectionPeriod.html', context)
@login_required(login_url="login")

def view_dotthu_detail_modal(request, pk):
    dot_thu = get_object_or_404(DotThuPhi, id_dotthu=pk)
    danh_sach_hoa_don = dot_thu.hoa_dons.select_related('id_hokhau').all().order_by('id_hokhau__so_can_ho')
    ids_da_co = danh_sach_hoa_don.values_list('id_hokhau_id', flat=True)
    tat_ca_ho_khau = HoKhau.objects.exclude(id_hokhau__in=ids_da_co)
    context = {
        'dot_thu': dot_thu,
        'danh_sach_hoa_don': danh_sach_hoa_don,
        'tat_ca_ho_khau': tat_ca_ho_khau # Bi·∫øn n√†y d√πng ƒë·ªÉ hi·ªÉn th·ªã trong ph·∫ßn "Th√™m h·ªô"
    }
    return render(request, 'core/ViewPeriodDetailModal.html', context)
@login_required(login_url="login")

def update_payment_status(request):
    invoice_ids = request.POST.getlist('invoice_ids[]') # Nh·∫≠n danh s√°ch ID t·ª´ AJAX
    
    if not invoice_ids:
        return JsonResponse({'status': 'error', 'message': 'Kh√¥ng c√≥ h√≥a ƒë∆°n n√†o ƒë∆∞·ª£c ch·ªçn'}, status=400)
    
    try:
        # C·∫≠p nh·∫≠t ng√†y n·ªôp l√† ng√†y hi·ªán t·∫°i cho c√°c h√≥a ƒë∆°n ƒë∆∞·ª£c ch·ªçn
        HoaDon.objects.filter(id_hoadon__in=invoice_ids, ngay_nop__isnull=True).update(
            ngay_nop=timezone.now()
        )
        return JsonResponse({'status': 'success', 'message': 'C·∫≠p nh·∫≠t tr·∫°ng th√°i th√†nh c√¥ng'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
@login_required(login_url="login")

def create_invoices_for_period(request):
    id_dotthu = request.POST.get('id_dotthu')
    hokhau_ids = request.POST.getlist('hokhau_ids[]')
    multipliers = request.POST.getlist('multipliers[]')
    prices = request.POST.getlist('prices[]')
    
    dot_thu = get_object_or_404(DotThuPhi, id_dotthu=id_dotthu)
    
    # Logic g√°n ID th·ªß c√¥ng cho HoaDon
    last_invoice = HoaDon.objects.all().order_by('id_hoadon').last()
    next_id = (last_invoice.id_hoadon + 1) if last_invoice else 1

    for hk_id, mult, price in zip(hokhau_ids, multipliers, prices):
        hokhau = get_object_or_404(HoKhau, id_hokhau=hk_id)
        
        if not HoaDon.objects.filter(id_dotthu=dot_thu, id_hokhau=hokhau).exists():
            # T√≠nh to√°n: T·ªïng ti·ªÅn = ƒê∆°n gi√° x H·ªá s·ªë
            final_amount = float(price) * float(mult)
            
            HoaDon.objects.create(
                id_hoadon=next_id,
                id_dotthu=dot_thu,
                id_hokhau=hokhau,
                tong_tien=final_amount # L∆∞u k·∫øt qu·∫£ ƒë√£ nh√¢n
            )
            next_id += 1

    # Tr·∫£ v·ªÅ c√πng Template ƒë·ªÉ n·∫°p l·∫°i b·∫£ng b√™n d∆∞·ªõi (ph·∫ßn Ch·ªù thu)
    danh_sach_hoa_don = dot_thu.hoa_dons.select_related('id_hokhau').all().order_by('id_hokhau__so_can_ho')
    tat_ca_ho_khau = HoKhau.objects.exclude(id_hokhau__in=danh_sach_hoa_don.values_list('id_hokhau_id', flat=True))
    
    return render(request, 'core/ViewPeriodDetailModal.html', {
        'dot_thu': dot_thu,
        'danh_sach_hoa_don': danh_sach_hoa_don,
        'tat_ca_ho_khau': tat_ca_ho_khau
    })
@login_required(login_url="login")

def add_dotthu(request):
    if request.method == 'POST':
        form = DotThuPhiForm(request.POST)
        if form.is_valid():
            new_ID = form.cleaned_data.get('id_dotthu')
            if DotThuPhi.objects.filter(id_dotthu=new_ID).exists():
                return JsonResponse({'error': 'M√£ ƒë·ª£t thu ph√≠ n√†y ƒë√£ t·ªìn t·∫°i!'}, status=400)
            
            form.save()
            
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'success'})
            
            return redirect('fee_collection_period') 
    else:
        form = DotThuPhiForm()
    return render(request, 'core/AddPeriodModal.html', {'form': form})
@login_required(login_url="login")

def edit_dotthu(request, pk):
    dot_thu = get_object_or_404(DotThuPhi, id_dotthu=pk) 
    if request.method == 'POST':
        form = DotThuPhiForm(request.POST, instance=dot_thu)
        if form.is_valid():
            form.save() 
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'status': 'success'})
            return redirect('fee_collection_period')
        else:
            return render(request, 'core/EditPeriodModal.html', {'form': form, 'dot_thu': dot_thu}, status=400)
            
    form = DotThuPhiForm(instance=dot_thu)
    return render(request, 'core/EditPeriodModal.html', {'form': form, 'dot_thu': dot_thu})
@login_required(login_url="login")

def delete_dotthu(request, pk):
    dot_thu = get_object_or_404(DotThuPhi, id_dotthu=pk)
    if request.method == 'POST':
        dot_thu.delete()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'status': 'success', 'message': 'ƒê·ª£t thu ph√≠ ƒë√£ ƒë∆∞·ª£c x√≥a th√†nh c√¥ng!'})
        return redirect('fee_collection_period')
    return render(request, 'core/DeletePeriodModal.html', {'dot_thu': dot_thu})
@login_required(login_url="login")

#Th·ªëng k√™
def statistics_view(request):
    # L·∫•y nƒÉm t·ª´ tham s·ªë GET, m·∫∑c ƒë·ªãnh l√† nƒÉm hi·ªán t·∫°i
    year_str = request.GET.get('year', str(datetime.now().year))
    try:
        year = int(year_str)
    except ValueError:
        year = datetime.now().year

    # 1. Bi·ªÉu ƒë·ªì tr√≤n: C∆° c·∫•u lo·∫°i kho·∫£n thu theo doanh thu trong NƒÇM ƒê√É CH·ªåN
    phi_status = KhoanThu.objects.filter(
        dot_thuphi__hoa_dons__ngay_nop__year=year
    ).annotate(
        total_revenue=Sum('dot_thuphi__hoa_dons__tong_tien')
    ).order_by('-total_revenue')[:5]
    
    pie_labels = [item.ten_khoanthu for item in phi_status]
    pie_data = [float(item.total_revenue) for item in phi_status]

    # 2. Bi·ªÉu ƒë·ªì ƒë∆∞·ªùng: Xu h∆∞·ªõng doanh thu theo th√°ng
    monthly_income = HoaDon.objects.filter(ngay_nop__year=year).annotate(
        month=ExtractMonth('ngay_nop')
    ).values('month').annotate(total=Sum('tong_tien')).order_by('month')

    line_labels = ["T1", "T2", "T3", "T4", "T5", "T6", "T7", "T8", "T9", "T10", "T11", "T12"]
    line_data = [0] * 12 
    for item in monthly_income:
        if item['month']:
            line_data[item['month'] - 1] = float(item['total'])
    
    # 3. T√≠nh to√°n c√°c con s·ªë t·ªïng quan
    tong_da_thu = HoaDon.objects.filter(ngay_nop__year=year).aggregate(Sum('tong_tien'))['tong_tien__sum'] or 0
    
    so_ho = HoKhau.objects.filter(is_active=True).count()
    dot_thus = DotThuPhi.objects.filter(ngay_batdau__year=year)
    tong_can_thu = 0
    for dot in dot_thus:
        # M·ª•c ti√™u d·ª± ki·∫øn = ƒê∆°n gi√° x S·ªë h·ªô (ƒê∆°n gi·∫£n h√≥a)
        tong_can_thu += float(dot.id_khoanthu.don_gia) * so_ho

    phan_tram = round((float(tong_da_thu) / tong_can_thu) * 100, 1) if tong_can_thu > 0 else 0

    context = {
        'pie_labels': json.dumps(pie_labels),
        'pie_data': json.dumps(pie_data),
        'line_labels': json.dumps(line_labels),
        'line_data': json.dumps(line_data),
        'tong_da_thu': tong_da_thu,
        'tong_can_thu': tong_can_thu,
        'phan_tram': phan_tram,
        'selected_year': year,
    }
    return render(request, 'core/Statistics.html', context)
@login_required(login_url="login")

def export_finance_excel(request):
    year_str = request.GET.get('year', str(timezone.now().year))
    try:
        year = int(year_str)
    except ValueError:
        year = timezone.now().year
        
    wb = Workbook()
    ws = wb.active
    ws.title = f"Thong ke {year}"

    # ƒê·ªãnh d·∫°ng
    bold_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    # S·ª¨A T·∫†I ƒê√ÇY: D√πng PatternFill tr·ª±c ti·∫øp thay v√¨ openpyxl.styles.PatternFill
    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")

    # Ti√™u ƒë·ªÅ b√°o c√°o
    ws.merge_cells('A1:C1')
    ws['A1'] = f"B√ÅO C√ÅO T√ÄI CH√çNH NƒÇM {year} - CHUNG C∆Ø BLUEMOON"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_align

    # Header b·∫£ng
    headers = ['Th√°ng', 'S·ªë l∆∞·ª£ng h√≥a ƒë∆°n', 'Doanh thu th·ª±c t·∫ø (VND)']
    ws.append([]) # D√≤ng tr·ªëng
    ws.append(headers)
    
    # ƒê·ªãnh d·∫°ng d√≤ng Header (D√≤ng s·ªë 3)
    for cell in ws[3]:
        cell.font = bold_font
        cell.fill = header_fill # S·ª≠ d·ª•ng bi·∫øn header_fill ƒë√£ t·∫°o
        cell.alignment = center_align
        cell.border = border

    # L·∫•y d·ªØ li·ªáu v√† ƒëi·ªÅn v√†o b·∫£ng
    monthly_data = HoaDon.objects.filter(ngay_nop__year=year).annotate(
        month=ExtractMonth('ngay_nop')
    ).values('month').annotate(
        count=Count('id_hoadon'),
        total=Sum('tong_tien')
    ).order_by('month')

    total_year_money = 0
    for m in range(1, 13):
        data = next((item for item in monthly_data if item['month'] == m), None)
        count = data['count'] if data else 0
        money = float(data['total']) if data else 0
        total_year_money += money
        
        ws.append([f"Th√°ng {m}", count, money])
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = Alignment(horizontal='right') if isinstance(cell.value, (int, float)) else center_align

    # D√≤ng t·ªïng c·ªông
    ws.append(["T·ªîNG C·ªòNG", "", total_year_money])
    last_row = ws.max_row
    ws.cell(row=last_row, column=1).font = Font(bold=True)
    ws.cell(row=last_row, column=3).font = Font(bold=True)
    for cell in ws[last_row]:
        cell.border = border

    # Ch·ªânh ƒë·ªô r·ªông c·ªôt
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=BaoCao_BlueMoon_{year}.xlsx'
    wb.save(response)
    return response